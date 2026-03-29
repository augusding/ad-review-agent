"""
审核报告导出模块。

支持 Excel 和 PDF 两种格式导出审核记录和统计数据。
"""
import io
import json
from datetime import datetime

from loguru import logger


def generate_excel_report(records: list[dict], stats: dict) -> bytes:
    """
    生成 Excel 报告，包含审核汇总和明细两个 Sheet。

    Args:
        records: 审核记录列表（来自 /records API）
        stats: 统计数据（来自 /stats API）

    Returns:
        Excel 文件的字节内容
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()

    # ── Sheet 1: 审核汇总 ──
    ws1 = wb.active
    ws1.title = "审核汇总"

    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws1["A1"] = "广告素材合规审核报告"
    ws1["A1"].font = Font(bold=True, size=18)
    ws1["A2"] = f"生成时间：{datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
    ws1["A3"] = f"记录数：{len(records)} 条"

    # 核心指标表格
    ws1["A5"] = "核心指标"
    ws1["A5"].font = header_font

    dist = stats.get("verdict_distribution", {})
    total = stats.get("total_reviews", 0)
    auto = dist.get("pass", 0) + dist.get("reject", 0)
    auto_rate = f"{auto / total * 100:.1f}%" if total > 0 else "-"

    metrics = [
        ("总审核量", str(total)),
        ("自动通过", str(dist.get("pass", 0))),
        ("自动拒绝", str(dist.get("reject", 0))),
        ("人工复核", str(dist.get("review", 0))),
        ("自动化率", auto_rate),
        ("平均置信度", f"{stats.get('avg_confidence', 0):.4f}"),
        ("平均处理时长", f"{stats.get('avg_processing_ms', 0)} ms"),
        ("降级次数", str(stats.get("fallback_count", 0))),
    ]

    for i, (label, value) in enumerate(metrics, start=6):
        ws1[f"A{i}"] = label
        ws1[f"A{i}"].font = label_font
        ws1[f"A{i}"].border = thin_border
        ws1[f"B{i}"] = value
        ws1[f"B{i}"].font = value_font
        ws1[f"B{i}"].border = thin_border
        ws1[f"B{i}"].alignment = Alignment(horizontal="right")

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: 审核明细 ──
    ws2 = wb.create_sheet("审核明细")

    columns = [
        "序号", "请求ID", "广告主ID", "品类", "素材类型",
        "Agent结论", "置信度", "违规数量", "主要违规维度",
        "人工结论", "处理时长(ms)", "提交时间",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    reject_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(records, start=2):
        v_count = r.get("violation_count", 0)
        v_dims = "-"
        if r.get("violations_json"):
            try:
                violations = json.loads(r["violations_json"])
                v_dims = ", ".join(set(v.get("dimension", "") for v in violations)) if violations else "-"
                v_count = v_count or len(violations)
            except (json.JSONDecodeError, TypeError):
                pass

        row_data = [
            row_idx - 1,
            r.get("request_id", ""),
            r.get("advertiser_id", ""),
            r.get("ad_category", ""),
            r.get("creative_type", ""),
            r.get("verdict", ""),
            r.get("confidence", 0),
            v_count,
            v_dims,
            r.get("human_verdict", "") or "-",
            r.get("processing_ms", 0),
            (r.get("created_at", "") or "")[:19],
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 7 and isinstance(val, float):
                cell.number_format = "0.00"

        verdict = r.get("verdict", "")
        if verdict == "pass":
            for col_idx in range(1, len(columns) + 1):
                ws2.cell(row=row_idx, column=col_idx).fill = pass_fill
        elif verdict == "reject" or v_count > 0:
            for col_idx in range(1, len(columns) + 1):
                ws2.cell(row=row_idx, column=col_idx).fill = reject_fill

    col_widths = [6, 28, 16, 12, 12, 10, 10, 10, 20, 10, 14, 20]
    for i, w in enumerate(col_widths, start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    ws2.auto_filter.ref = f"A1:L{len(records) + 1}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_report(records: list[dict], stats: dict) -> bytes:
    """
    生成 PDF 报告，包含封面、指标摘要和审核明细。

    Args:
        records: 审核记录列表
        stats: 统计数据

    Returns:
        PDF 文件的字节内容
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        PageBreak,
    )

    # 注册中文字体
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    cn_style = ParagraphStyle(
        "Chinese",
        parent=styles["Normal"],
        fontName="STSong-Light",
        fontSize=10,
        leading=14,
    )
    cn_title = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName="STSong-Light",
        fontSize=24,
        leading=30,
    )
    cn_h2 = ParagraphStyle(
        "ChineseH2",
        parent=styles["Heading2"],
        fontName="STSong-Light",
        fontSize=14,
        leading=20,
    )

    elements = []

    # ── 封面 ──
    elements.append(Spacer(1, 60 * mm))
    elements.append(Paragraph("广告素材合规审核报告", cn_title))
    elements.append(Spacer(1, 15 * mm))
    elements.append(Paragraph(
        f"生成时间：{datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", cn_style
    ))
    elements.append(Paragraph(f"记录数：{len(records)} 条", cn_style))
    elements.append(Paragraph("系统：Ad Review Agent - vivo 商业产品部", cn_style))
    elements.append(PageBreak())

    # ── 核心指标摘要 ──
    elements.append(Paragraph("核心指标摘要", cn_h2))
    elements.append(Spacer(1, 5 * mm))

    dist = stats.get("verdict_distribution", {})
    total = stats.get("total_reviews", 0)
    auto = dist.get("pass", 0) + dist.get("reject", 0)
    auto_rate = f"{auto / total * 100:.1f}%" if total > 0 else "-"

    metric_data = [
        ["指标", "值"],
        ["总审核量", str(total)],
        ["自动通过", str(dist.get("pass", 0))],
        ["自动拒绝", str(dist.get("reject", 0))],
        ["人工复核", str(dist.get("review", 0))],
        ["自动化率", auto_rate],
        ["平均置信度", f"{stats.get('avg_confidence', 0):.4f}"],
        ["平均处理时长", f"{stats.get('avg_processing_ms', 0)} ms"],
        ["降级次数", str(stats.get("fallback_count", 0))],
        ["队列待处理", str(stats.get("queue_pending", 0))],
    ]

    t = Table(metric_data, colWidths=[60 * mm, 50 * mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    elements.append(PageBreak())

    # ── 审核明细 ──
    elements.append(Paragraph("审核明细", cn_h2))
    elements.append(Spacer(1, 5 * mm))

    header = ["#", "请求ID", "品类", "结论", "置信度", "违规数", "耗时(ms)", "时间"]
    detail_data = [header]

    for i, r in enumerate(records, start=1):
        v_count = r.get("violation_count", 0)
        detail_data.append([
            str(i),
            r.get("request_id", "")[:20],
            r.get("ad_category", ""),
            r.get("verdict", ""),
            f"{r.get('confidence', 0):.2f}",
            str(v_count),
            str(r.get("processing_ms", 0)),
            (r.get("created_at", "") or "")[:10],
        ])

    col_w = [8 * mm, 38 * mm, 18 * mm, 14 * mm, 16 * mm, 14 * mm, 18 * mm, 22 * mm]
    dt = Table(detail_data, colWidths=col_w, repeatRows=1)

    style_cmds = [
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (4, 0), (6, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]

    # Color rows by verdict
    for i, r in enumerate(records, start=1):
        verdict = r.get("verdict", "")
        if verdict == "reject":
            style_cmds.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FCE4EC"))
            )
        elif verdict == "pass":
            style_cmds.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E2EFDA"))
            )

    dt.setStyle(TableStyle(style_cmds))
    elements.append(dt)

    doc.build(elements)
    return output.getvalue()
